from qa import QA
from user import get_user
from quiz import Quiz, Question
from sys import argv
from random import shuffle
from openpyxl import load_workbook

# add a QA with the tags entered as args 

def add_qa(user):
    try:                                                  
        q = argv[2]
        a = argv[3]
    except:
        print('Add at least two parameters (Q and A).')
        exit()
    
    tags = argv[4:]
    qa = QA(q, a, tags)
    user.add_qa(qa)

# add all QA from the .xlsx (1st and 2nd columns) with the tags entered as args

def add_qa_from_xlsx(user):
    try:
        filename = argv[2]
    except:
        print('Argument missing (.xlsx filename).')
        exit()

    tags = argv[3:]

    wb = load_workbook(filename = filename)
    ws = wb.active
    length = len(list(ws.rows))
    lines = [(ws.cell(i, 1).value, ws.cell(i, 2).value) for i in range(1,20)]
    
    for line in lines:
        q = line[0]
        a = line[1]
        qa = QA(q, a, tags)
        user.add_qa(qa)

# display all QA related to the tags entered as args

def display_qa_list(user):
    tags = argv[2:]

    user.display_qa_list(tags = tags)

# generate a quiz considering the tags entered as args

def generate_quiz(user, reverse = False):
    tags = argv[2:]

    qa_list = user.generate_qa_list(tags = tags)
    if qa_list == []:
        print('No Q&A available.')
        exit()

    shuffle(qa_list)

    question_list = [Question(qa[1], qa[2], qa[3]) for qa in qa_list]

    quiz = Quiz(question_list, reverse)
    quiz.run()

# display all tags or the tags related to the ones entered as args

def display_tags(user):
    tags = argv[2:]
    all_tags = user.get_tags(tags = tags)
    print(all_tags)

def remove(user):
    pass

def switch_user():
    pass


if __name__ == '__main__':
    
    user = get_user()
    
    try:
        arg1 = argv[1]
    except:
        print('Missing parameters.')
        exit()

    match arg1:
        case '-a':
            add_qa(user)
        case '-ax':
            add_qa_from_xlsx(user)            
        case '-d':
            display_qa_list(user)
        case '-q':
            generate_quiz(user)
        case '-qr':
            generate_quiz(user, reverse = True)
        case '-t':
            display_tags(user)
        case '-rm':
            remove(user)
        case '-su':
            switch_user()
        case _:
            print('The first parameter is not appropriate.')
